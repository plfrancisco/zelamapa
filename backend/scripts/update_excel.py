import openpyxl

file_path = '/home/lucaspedro81/Área de trabalho/Projeto Integrador /ZelaMapa/Diário de Bordo - ZelaMapa - PI-II-26-1.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Diário de bordo'] if 'Diário de bordo' in wb.sheetnames else wb.active

data_to_fill = {
    14: { # Semana 9
        4: "Estruturação inicial da arquitetura do repositório.", # D: Planejadas
        5: "Criação da base estrutural com integração GovTech e pacotes React-Leaflet.", # E: Realizadas
        6: "Pedro / Grupo.", # F: Responsáveis
        7: "Adaptação ao formato do repositório monorepo ou estrutura inicial.", # G: Dificuldades
        8: "Definição clara das dependências e scripts de inicialização (iniciar_projeto).", # H: Soluções
        9: "Base de pastas e dependências concluída.", # I: Resultados
        10: "Início da documentação dos diagramas.", # J: Próximos passos
        11: "Commits iniciais." # K: Evidências
    },
    15: { # Semana 10
        4: "Elaboração de diagramas arquiteturais e documentação.",
        5: "Criação de representação de diagramas MER, DFD e Casos de Uso.",
        6: "Grupo todo.",
        7: "Formalizar os diagramas de maneira simples para a documentação textual.",
        8: "Uso de notação textual simplificada direto no repositório.",
        9: "Documentação base estabelecida.",
        10: "Testes locais e avanço inicial das telas.",
        11: "Atualização README e docs."
    },
    16: { # Semana 11
        4: "Integração do mapa e funções de formulário.",
        5: "Implementação das bibliotecas Leaflet para mapas dinâmicos e OSRM/ViaCEP.",
        6: "Pedro Lucas Francisco.",
        7: "Substituição de protótipos estáticos do Figma por componentes dinâmicos.",
        8: "Uso de bibliotecas consolidadas como react-leaflet para evitar retrabalho.",
        9: "Mapa dinâmico no painel funcional.",
        10: "Estruturar o board do Trello para alinhar entrega final.",
        11: "Commits de components/map."
    },
    17: { # Semana 12
        4: "Organização das tarefas e preparação para Deploy.",
        5: "Criação do board no Trello com cards de plano de Deployment.",
        6: "Grupo todo.",
        7: "Orquestrar passos sistemáticos de deploy sem quebrar dependências.",
        8: "Divisão do processo por etapas (Frontend, Backend, Monitoramento).",
        9: "Pipeline de tarefas na nuvem clara e definida.",
        10: "Implementar tracking em tempo real.",
        11: "Board Trello atualizado."
    },
    18: { # Semana 13
        4: "Desenvolvimento do rastreamento em tempo real.",
        5: "Implementação de real-time tracking estilo Waze e painel de inteligência.",
        6: "Pedro Lucas Francisco.",
        7: "Sincronização do frontend com eventos do socket/servidor de simulação.",
        8: "Refatoração de arquivos e limpeza de dependências do planejamento.",
        9: "Navegação e visualização de rotas funcionais.",
        10: "Início do Deploy para produção e Vercel.",
        11: "Commits de 'real-time tracking'."
    },
    19: { # Semana 14 (up to 19/4)
        4: "Deploy do pacote completo e correções de conflitos.",
        5: "Deploy na Vercel, correção da versão Vite 8 removendo PWA, finalização do branding 'ZelaMapa'.",
        6: "Pedro Lucas Francisco.",
        7: "Erros de build na Vercel (incompatibilidade do plugin pwa).",
        8: "Remoção do vite-plugin-pwa, clear de node_modules (clean lock).",
        9: "Deploy final estabilizado, repositório limpo.",
        10: "Apresentação e testes com avaliadores.",
        11: "Link Vercel / Repositório limpo."
    }
}

for row_idx, cols in data_to_fill.items():
    for col_idx, value in cols.items():
        if ws.cell(row=row_idx, column=col_idx).value is None or str(ws.cell(row=row_idx, column=col_idx).value).strip() == "":
            ws.cell(row=row_idx, column=col_idx).value = value

wb.save(file_path)
print("Excel atualizado com sucesso!")
